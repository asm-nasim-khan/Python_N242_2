# import json

# x = '{"name" : "Nasim", "age":99}'

# # print(type(x))
# y = json.loads(x)
# # print(type(y))

# print(y['age'])

# myDict = {1: 'One', 2: 'Two', 3: 'Three', 4: 'Four', 5: 'Five'}
# my_json = json.dumps(myDict)
# # print(type(my_json))
# # print(my_json)
# z = json.loads(my_json)
# print(type(z))

# my_list = [1,2,3]

# iterator_1 = iter(my_list)
# print(next(iterator_1))
# print(next(iterator_1))
# print(next(iterator_1))
# # print(next(iterator_1))

# word = "Mango"
# # Iterator_2 = iter(word)
# # print(next(Iterator_2))
# # print(next(Iterator_2))
# for cha in word:
#     print(cha)


#  Generators

# def NUM_Gen():
#     yield 20
#     yield 29
#     yield 28

# iterator_3 = NUM_Gen()
# print(next(iterator_3))
# print(next(iterator_3))
# print(next(iterator_3))

# for item in NUM_Gen():
#     print(item)


# Decorators
# def MakeUpper(text):
#     return text.upper()

# Decorators_1 = MakeUpper
# print(Decorators_1("how are you"))


#  Excel File – Read, Write, Append, Arithmetic Operation 

# Method_1 

# import pandas as pd
# dataFrame_1 = pd.read_excel('PYTH-Batch-N242-File.xlsx')
# print(dataFrame_1['Class'])

# Method_2:

import openpyxl as oxl

# excel_file = oxl.load_workbook('PYTH-Batch-N242-File.xlsx')
# excel_file = excel_file.active
# print(excel_file.max_row)
# print(excel_file.max_column)

# for row in range(0,excel_file.max_row):
# for row in range(1,50):
#     # for col in excel_file.insert_cols(1,excel_file.max_column):
#     for col in excel_file.iter_cols(1,4):
#         print(col[row].value,end= "||")
#     print()
Ex_file  = oxl.Workbook()
sheet = Ex_file.active
sheet.title = "Class_24"
# c1 = sheet.cell(row=1,column=1)
# c1.value = "Eshikhon"

for row_x in range(1,50):
    for col_x in range(1,50):
        if row_x == col_x:
            temp = sheet.cell(row=row_x,column=col_x)
            temp.value = "Python"
        else:
            temp = sheet.cell(row=row_x,column=col_x)
            temp.value = "Eshikhon"

# Ex_file.save("NEW_Generated_File_Class_24.xlsx") #To save the file

print(sheet["B6"].value + sheet["A1"].value)


